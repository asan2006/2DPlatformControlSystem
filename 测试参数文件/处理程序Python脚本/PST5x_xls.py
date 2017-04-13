import os
import sys
import json
import openpyxl
import time
from openpyxl import load_workbook

# 改变表头的值
header = {
    'A1':'Time(S)',
    'B1':'ADXL335_In_X',
    'C1':'ADXL335_In_Y',
    'D1':'ADXL335_Ex_Y',
    'E1':'LPR403AL',
    'F1':'ADXL335_In_X_Ave',
    'G1':'ADXL335_In_Y_Ave',
    'H1':'ADXL335_Ex_Y_Ave',
    'I1':'LPR403AL_Ave'
    }
sheetOriginalName = 'Sheet2'

def main():
    path = os.getcwd()
    files = os.listdir()
    for f in files:
        _,ext = os.path.splitext(f)
        if ext in ('.xls','.xlsx'):
            PST5x_xls(f)

def PST5x_xls(xlsFile):
    wb = load_workbook(xlsFile)
    ws = wb.create_sheet(title='Decoder')
    # 判断并处理feedback.txt文件
    feedbackFileName = os.path.splitext(xlsFile)[0] + '_Feedback.txt'
    if os.path.exists(feedbackFileName):
        n = 0
        fo = open(feedbackFileName)
        data = fo.readlines()
        for l in data:
            n+=1
            print('正在写入反馈数据第{0}行'.format(n),end='\r')
            ws.append(l.rsplit())
        #for line in open(feedbackFileName):
        #    n+=1
        #    print('正在写入反馈数据第{0}行'.format(n),end='\r')
        #    ws.append(line.rsplit())
        print('{0}写入反馈数据成功'.format(xlsFile))
        wb.save(xlsFile)
        print('保存{0}成功'.format(xlsFile))

def test():
    for i in range(301):
        print('processing %d out od %d items...' % (i + 1,301),end='\r')
        time.sleep(0.1)
        
if __name__ == "__main__":
    main()
    input()
