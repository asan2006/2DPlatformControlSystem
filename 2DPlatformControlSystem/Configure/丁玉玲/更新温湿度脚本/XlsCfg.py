import sys
import os
import openpyxl

PATHFILES = "文件路径.txt"
T_RH_FILE = "温湿度信息.txt"
N = 0


def readNew_T_RH():
    f = open(T_RH_FILE)
    newInfo = f.read()
    f.close()
    return newInfo

def UpdateXlsCfg_T_RH(xlsFile):
    print("正在更新第{0}个配置文件：{1}".format(N,xlsFile))
    wb = openpyxl.load_workbook(xlsFile)
    ws = wb.worksheets[0]
    expInfo = str(ws['B27'].value)  
    expLines = expInfo.splitlines()
    #更新最后一行温湿度信息
    expLines[-1] = readNew_T_RH()
    newExpInfo = ("\n").join(expLines)
    #更新保存配置文件
    ws['B27'] = newExpInfo
    ws['B19'] = 0
    ws['B20'] = 0
    ws['B21'] = 0
    ws['B22'] = 0
    wb.save(xlsFile)
    print("更新第{0}个成功：{1}".format(N,xlsFile))

def main():
    global N
    #读取文件夹路径
    filePaths = []
    for line in open("文件路径.txt"):
        filePaths.append(line.strip("\n"))
        #print(filePaths)

    #读取每一个配置文件
    for path in filePaths:    
        listFile = os.listdir(path)
        for file in listFile:
            if os.path.isfile(path + "\\" + file):
                # 判断是xls类型的配置文件
                if os.path.splitext(file)[1] in [".xlsx",".xls"]:
                    N = N + 1
                    UpdateXlsCfg_T_RH(path + "\\" + file)

if __name__ == "__main__":
    main()

